# from mypack.m1 import add
# from mypack.m2 import sub
# print(add(5, 3))
# print(sub(3, 5))
import openpyxl as excel


# wb = excel.Workbook()
# ws = wb.active
# ws['A1'] = "HELLO"
# ws['A2'] = 24
# ws['A3'] = "chill out"
# wb.save('chill.xlsx')
# wb = excel.load_workbook(filename='chill.xlsx')
# sheet = wb['mysheet']

# for row in sheet.iter_rows(min_row = 1, max_row = sheet.max_row,min_col = 1, max_col = sheet.max_column):
#     for cell in row: 
#         print(cell.value, end=" ")
#     print()


# from scipy.optimize import minimize
# import numpy as np

# def fun(x):
#     return x ** 2 + 10 * np.sin(x)
# res = minimize(fun, 1.9)
# print(res.x)